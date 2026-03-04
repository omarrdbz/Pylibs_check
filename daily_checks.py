"""
daily_checks.py — Verificador diario de versiones y vulnerabilidades
====================================================================
Lee un Excel con las librerías Python aprobadas en la empresa,
consulta PyPI (últimas versiones) y OSV.dev (vulnerabilidades),
y genera un reporte HTML listo para publicar en Microsoft Teams.

Uso:
    python daily_checks.py --input libs.xlsx
    python daily_checks.py --input libs.xlsx --output reporte.html
"""

import argparse
import json
import logging
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from logging.handlers import RotatingFileHandler
from typing import Any

import requests
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook
from packaging.version import InvalidVersion, Version

# ============================================================================
# Constantes
# ============================================================================
PYPI_BASE_URL = "https://pypi.org/pypi"
OSV_BATCH_URL = "https://api.osv.dev/v1/querybatch"
OSV_VULN_URL = "https://api.osv.dev/v1/vulns"

HTTP_TIMEOUT = 10           # segundos por request
OSV_TIMEOUT = 30            # segundos para el batch de OSV
RATE_LIMIT_DELAY = 0.2      # segundos entre requests a PyPI
MAX_RETRIES = 3
BACKOFF_FACTOR = 0.5
OSV_BATCH_SIZE = 1000       # máximo de queries por batch de OSV
RECENT_VULNS_HOURS = 24     # período para buscar vulnerabilidades recientes

LOG_FILE = "daily_checks.log"
LOG_MAX_BYTES = 5_242_880   # 5 MB
LOG_BACKUP_COUNT = 3

TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")
TEMPLATE_FILE = "report_template.html"

# Columnas esperadas en el Excel (case-insensitive)
COL_LIBRARY = "Library"
COL_VERSION = "Version"
COL_APPROVED = "Approved"    # opcional: si existe, solo procesa filas aprobadas
COL_LANGUAGE = "Language"    # opcional: si existe, solo procesa filas con 'python'


# ============================================================================
# Modelos de datos
# ============================================================================
@dataclass
class VulnerabilityInfo:
    """Representa una vulnerabilidad individual detectada por OSV."""
    id: str
    aliases: list = field(default_factory=list)
    summary: str = ""
    severity: str = "UNKNOWN"
    fixed_version: str = ""


@dataclass
class RecentVuln:
    """Vulnerabilidad publicada/modificada recientemente (últimas 24h)."""
    id: str
    library: str
    aliases: list = field(default_factory=list)
    summary: str = ""
    severity: str = "UNKNOWN"
    fixed_version: str = ""
    modified: str = ""            # fecha ISO de última modificación
    affected_versions: str = ""   # rango de versiones afectadas


@dataclass
class LibraryResult:
    """Resultado consolidado de la verificación de una librería."""
    name: str
    current_version: str
    latest_version: str = ""
    version_status: str = "OK"    # OK | UPDATE | ERROR (solo versión, no cambia)
    status: str = "OK"            # Estado compuesto para el badge visual
    pypi_url: str = ""
    vulnerabilities: list = field(default_factory=list)
    error_message: str = ""


@dataclass
class ReportStats:
    """Estadísticas agregadas para el reporte."""
    total: int = 0
    outdated: int = 0
    vulnerable: int = 0
    errors: int = 0
    recent_vulns: int = 0


# ============================================================================
# Logging
# ============================================================================
def setup_logging() -> logging.Logger:
    """Configura logging con rotación de archivos y salida a consola."""
    logger = logging.getLogger("daily_checks")
    logger.setLevel(logging.INFO)

    formatter = logging.Formatter(
        "[%(asctime)s] %(levelname)-8s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    # Archivo con rotación
    try:
        fh = RotatingFileHandler(
            LOG_FILE, maxBytes=LOG_MAX_BYTES, backupCount=LOG_BACKUP_COUNT, encoding="utf-8"
        )
        fh.setFormatter(formatter)
        logger.addHandler(fh)
    except (OSError, PermissionError) as exc:
        print(f"[WARN] No se pudo crear '{LOG_FILE}': {exc}")

    # Consola
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(formatter)
    logger.addHandler(ch)

    return logger


# ============================================================================
# Sesión HTTP con reintentos
# ============================================================================
def create_http_session() -> requests.Session:
    """Crea sesión HTTP con retry automático en errores transitorios."""
    session = requests.Session()
    retry = requests.packages.urllib3.util.retry.Retry(
        total=MAX_RETRIES,
        backoff_factor=BACKOFF_FACTOR,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "POST"],
    )
    adapter = requests.adapters.HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update({"Accept": "application/json", "User-Agent": "daily-checks/1.0"})
    return session


# ============================================================================
# Lector de Excel
# ============================================================================
def read_approved_libraries(file_path: str, logger: logging.Logger) -> list[dict]:
    """
    Lee el Excel y retorna lista de dicts {'name': ..., 'version': ...}.
    Filtra por columna 'Approved' (si existe) y 'Language' (si existe, solo Python).
    """
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"No se encontró el archivo: {file_path}")

    logger.info("Leyendo archivo Excel: %s", file_path)
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    if ws is None:
        raise ValueError("El archivo Excel no contiene hojas activas.")

    # Leer encabezados (primera fila)
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers_lower = [str(h).strip().lower() if h else "" for h in headers]

    # Buscar columnas requeridas (case-insensitive)
    try:
        idx_name = headers_lower.index(COL_LIBRARY.lower())
    except ValueError:
        raise ValueError(f"Columna '{COL_LIBRARY}' no encontrada. Columnas disponibles: {headers}")

    try:
        idx_version = headers_lower.index(COL_VERSION.lower())
    except ValueError:
        raise ValueError(f"Columna '{COL_VERSION}' no encontrada. Columnas disponibles: {headers}")

    # Columnas opcionales de filtrado
    idx_approved = None
    try:
        idx_approved = headers_lower.index(COL_APPROVED.lower())
        logger.info("Columna '%s' detectada — se filtrarán solo librerías aprobadas.", COL_APPROVED)
    except ValueError:
        pass

    idx_language = None
    try:
        idx_language = headers_lower.index(COL_LANGUAGE.lower())
        logger.info("Columna '%s' detectada — se filtrarán solo librerías Python.", COL_LANGUAGE)
    except ValueError:
        pass

    # Leer filas
    libraries = []
    skipped = 0
    skipped_not_approved = 0
    skipped_not_python = 0
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = row[idx_name] if idx_name < len(row) else None
        version = row[idx_version] if idx_version < len(row) else None

        if not name or not str(name).strip():
            skipped += 1
            continue
        if not version or not str(version).strip():
            logger.warning("Fila %d: versión vacía para '%s', se omite.", row_num, name)
            skipped += 1
            continue

        # Filtro: columna Approved
        if idx_approved is not None:
            approved_val = row[idx_approved] if idx_approved < len(row) else None
            if not _is_truthy(approved_val):
                skipped_not_approved += 1
                continue

        # Filtro: columna Language
        if idx_language is not None:
            lang_val = row[idx_language] if idx_language < len(row) else None
            if not lang_val or str(lang_val).strip().lower() != "python":
                skipped_not_python += 1
                continue

        libraries.append({"name": str(name).strip(), "version": str(version).strip()})

    wb.close()

    logger.info("Se leyeron %d librerías Python aprobadas.", len(libraries))
    if skipped:
        logger.info("  Filas omitidas (datos incompletos): %d", skipped)
    if skipped_not_approved:
        logger.info("  Filas omitidas (no aprobadas):      %d", skipped_not_approved)
    if skipped_not_python:
        logger.info("  Filas omitidas (no Python):         %d", skipped_not_python)

    if not libraries:
        raise ValueError("El archivo Excel no contiene librerías Python aprobadas.")

    return libraries


def _is_truthy(value) -> bool:
    """Determina si un valor de celda es 'verdadero' (True, Yes, Sí, 1, X, etc.)."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    return text in ("true", "yes", "sí", "si", "1", "x", "y")


# ============================================================================
# Verificador de versiones (PyPI)
# ============================================================================
def check_latest_version(name: str, session: requests.Session, logger: logging.Logger) -> dict:
    """Consulta PyPI para la última versión. Retorna dict con latest_version, pypi_url, error."""
    url = f"{PYPI_BASE_URL}/{name}/json"
    try:
        resp = session.get(url, timeout=HTTP_TIMEOUT)
        resp.raise_for_status()
        info = resp.json().get("info", {})
        return {
            "latest_version": info.get("version", ""),
            "pypi_url": info.get("package_url", f"https://pypi.org/project/{name}/"),
            "error": "",
        }
    except requests.exceptions.HTTPError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            logger.warning("Paquete '%s' no encontrado en PyPI.", name)
            return {"latest_version": "", "pypi_url": "", "error": "No encontrado en PyPI"}
        logger.error("Error HTTP para '%s': %s", name, exc)
        return {"latest_version": "", "pypi_url": "", "error": str(exc)}
    except requests.exceptions.RequestException as exc:
        logger.error("Error de red para '%s': %s", name, exc)
        return {"latest_version": "", "pypi_url": "", "error": str(exc)}


def compare_versions(current: str, latest: str) -> str:
    """Compara versiones con PEP 440. Retorna 'OK' o 'UPDATE'."""
    try:
        return "OK" if Version(current) >= Version(latest) else "UPDATE"
    except InvalidVersion:
        return "OK" if current.strip() == latest.strip() else "UPDATE"


def check_all_versions(libraries: list[dict], session: requests.Session,
                       logger: logging.Logger) -> list[LibraryResult]:
    """Consulta PyPI para todas las librerías con rate limiting."""
    results = []
    total = len(libraries)

    for idx, lib in enumerate(libraries, start=1):
        logger.info("[%d/%d] Consultando PyPI: %s==%s", idx, total, lib["name"], lib["version"])

        pypi = check_latest_version(lib["name"], session, logger)
        result = LibraryResult(
            name=lib["name"],
            current_version=lib["version"],
            latest_version=pypi["latest_version"],
            pypi_url=pypi["pypi_url"],
        )

        if pypi["error"]:
            result.version_status = "ERROR"
            result.status = "ERROR"
            result.error_message = pypi["error"]
        elif pypi["latest_version"]:
            ver_status = compare_versions(lib["version"], pypi["latest_version"])
            result.version_status = ver_status
            result.status = ver_status
        else:
            result.version_status = "ERROR"
            result.status = "ERROR"
            result.error_message = "No se pudo determinar la última versión"

        results.append(result)

        # Rate limiting (no esperar después del último)
        if idx < total:
            time.sleep(RATE_LIMIT_DELAY)

    return results


# ============================================================================
# Scanner de vulnerabilidades (OSV.dev)
# ============================================================================
def query_osv_batch(libraries: list[dict], session: requests.Session,
                    logger: logging.Logger) -> dict[str, list[dict]]:
    """
    Consulta OSV.dev en batch. Retorna {nombre: [vulns crudas]}.
    """
    queries = [
        {"package": {"name": lib["name"], "ecosystem": "PyPI"}, "version": lib["version"]}
        for lib in libraries
    ]
    names = [lib["name"] for lib in libraries]
    results: dict[str, list[dict]] = {}

    logger.info("Consultando OSV.dev para %d librerías...", len(queries))

    # Enviar en batches de OSV_BATCH_SIZE
    for start in range(0, len(queries), OSV_BATCH_SIZE):
        batch_q = queries[start:start + OSV_BATCH_SIZE]
        batch_n = names[start:start + OSV_BATCH_SIZE]

        try:
            resp = session.post(OSV_BATCH_URL, json={"queries": batch_q}, timeout=OSV_TIMEOUT)
            resp.raise_for_status()

            for idx, vuln_result in enumerate(resp.json().get("results", [])):
                if idx >= len(batch_n):
                    break
                vulns = vuln_result.get("vulns", [])
                if vulns:
                    results[batch_n[idx]] = vulns
                    logger.warning("¡%d vulnerabilidad(es) en '%s'!", len(vulns), batch_n[idx])

        except requests.exceptions.RequestException as exc:
            logger.error("Error consultando OSV.dev: %s", exc)

    return results


def fetch_vuln_details(vuln_id: str, session: requests.Session,
                       logger: logging.Logger) -> VulnerabilityInfo:
    """Obtiene detalles de una vulnerabilidad por ID desde OSV.dev."""
    url = f"{OSV_VULN_URL}/{vuln_id}"
    try:
        resp = session.get(url, timeout=HTTP_TIMEOUT)
        resp.raise_for_status()
        data = resp.json()

        return VulnerabilityInfo(
            id=vuln_id,
            aliases=data.get("aliases", []),
            summary=data.get("summary", data.get("details", "Sin descripción"))[:200],
            severity=_extract_severity(data),
            fixed_version=_extract_fixed_version(data),
        )
    except requests.exceptions.RequestException as exc:
        logger.warning("No se pudieron obtener detalles de %s: %s", vuln_id, exc)
        return VulnerabilityInfo(id=vuln_id, summary="Error al obtener detalles")


def _extract_severity(data: dict) -> str:
    """Extrae severidad desde CVSS o database_specific."""
    # Desde campo severity (score CVSS)
    for sev in data.get("severity", []):
        score_str = sev.get("score", "")
        try:
            score = float(score_str)
            if score >= 9.0: return "CRITICAL"
            if score >= 7.0: return "HIGH"
            if score >= 4.0: return "MEDIUM"
            return "LOW"
        except ValueError:
            pass

    # Desde database_specific
    severity = data.get("database_specific", {}).get("severity", "")
    if isinstance(severity, str) and severity.upper() in ("CRITICAL", "HIGH", "MEDIUM", "LOW"):
        return severity.upper()

    return "UNKNOWN"


def _extract_fixed_version(data: dict) -> str:
    """Extrae la versión corregida desde affected[].ranges[].events[]."""
    for affected in data.get("affected", []):
        for rng in affected.get("ranges", []):
            for event in rng.get("events", []):
                fixed = event.get("fixed")
                if fixed:
                    return fixed
    return ""


def scan_vulnerabilities(libraries: list[dict], results: list[LibraryResult],
                         session: requests.Session, logger: logging.Logger) -> None:
    """
    Escanea vulnerabilidades de la versión aprobada y enriquece los LibraryResult in-place.
    """
    # Paso 1: batch query (con la versión aprobada del Excel)
    vulns_by_pkg = query_osv_batch(libraries, session, logger)
    if not vulns_by_pkg:
        logger.info("No se encontraron vulnerabilidades en las versiones aprobadas.")
        return

    # Paso 2: detalles individuales
    result_map = {r.name: r for r in results}

    for pkg_name, raw_vulns in vulns_by_pkg.items():
        if pkg_name not in result_map:
            continue

        lib_result = result_map[pkg_name]
        for raw_vuln in raw_vulns:
            vuln_id = raw_vuln.get("id", "")
            if not vuln_id:
                continue

            vuln_info = fetch_vuln_details(vuln_id, session, logger)
            lib_result.vulnerabilities.append(vuln_info)
            time.sleep(RATE_LIMIT_DELAY)

        if lib_result.vulnerabilities:
            lib_result.status = "VULNERABLE"


def scan_recent_vulnerabilities(libraries: list[dict], known_vuln_ids: set,
                                session: requests.Session,
                                logger: logging.Logger) -> list[RecentVuln]:
    """
    Busca vulnerabilidades publicadas/modificadas en las últimas 24 horas
    para las librerías del Excel, sin importar la versión.
    Excluye las que ya se reportaron en la sección de versión aprobada.
    """
    logger.info("Buscando vulnerabilidades recientes (últimas 24h)...")
    cutoff = datetime.now(timezone.utc) - timedelta(hours=RECENT_VULNS_HOURS)
    recent: list[RecentVuln] = []

    # Consulta OSV sin versión → retorna TODAS las vulns del paquete
    queries = [
        {"package": {"name": lib["name"], "ecosystem": "PyPI"}}
        for lib in libraries
    ]
    names = [lib["name"] for lib in libraries]

    for start in range(0, len(queries), OSV_BATCH_SIZE):
        batch_q = queries[start:start + OSV_BATCH_SIZE]
        batch_n = names[start:start + OSV_BATCH_SIZE]

        try:
            resp = session.post(OSV_BATCH_URL, json={"queries": batch_q}, timeout=OSV_TIMEOUT)
            resp.raise_for_status()

            for idx, vuln_result in enumerate(resp.json().get("results", [])):
                if idx >= len(batch_n):
                    break
                pkg_name = batch_n[idx]

                for raw_vuln in vuln_result.get("vulns", []):
                    vuln_id = raw_vuln.get("id", "")
                    modified_str = raw_vuln.get("modified", "")

                    # Saltar si ya se reportó en la sección de versión aprobada
                    if vuln_id in known_vuln_ids:
                        continue

                    # Filtrar por fecha de modificación (últimas 24h)
                    if not modified_str:
                        continue
                    try:
                        modified_dt = datetime.fromisoformat(modified_str.replace("Z", "+00:00"))
                        if modified_dt < cutoff:
                            continue
                    except ValueError:
                        continue

                    # Obtener detalles
                    detail = fetch_vuln_details(vuln_id, session, logger)

                    # Extraer rango de versiones afectadas desde el detalle
                    affected_str = _extract_affected_range(vuln_id, session, logger)

                    recent.append(RecentVuln(
                        id=vuln_id,
                        library=pkg_name,
                        aliases=detail.aliases,
                        summary=detail.summary,
                        severity=detail.severity,
                        fixed_version=detail.fixed_version,
                        modified=modified_dt.strftime("%Y-%m-%d %H:%M UTC"),
                        affected_versions=affected_str,
                    ))
                    time.sleep(RATE_LIMIT_DELAY)

        except requests.exceptions.RequestException as exc:
            logger.error("Error buscando vulns recientes: %s", exc)

    logger.info("Vulnerabilidades recientes encontradas: %d", len(recent))
    return recent


def _extract_affected_range(vuln_id: str, session: requests.Session,
                            logger: logging.Logger) -> str:
    """
    Obtiene el rango de versiones afectadas desde los detalles de OSV.
    Retorna algo como '>=1.0.0, <2.3.1' o vacío.
    """
    url = f"{OSV_VULN_URL}/{vuln_id}"
    try:
        resp = session.get(url, timeout=HTTP_TIMEOUT)
        resp.raise_for_status()
        data = resp.json()

        for affected in data.get("affected", []):
            ranges_parts = []
            for rng in affected.get("ranges", []):
                introduced = ""
                fixed = ""
                for event in rng.get("events", []):
                    if "introduced" in event:
                        introduced = event["introduced"]
                    if "fixed" in event:
                        fixed = event["fixed"]
                if introduced and fixed:
                    ranges_parts.append(f">={introduced}, <{fixed}")
                elif introduced:
                    ranges_parts.append(f">={introduced}")
                elif fixed:
                    ranges_parts.append(f"<{fixed}")
            if ranges_parts:
                return " | ".join(ranges_parts)
        return ""
    except Exception:
        return ""


# ============================================================================
# Generador de reporte HTML
# ============================================================================
def get_summary_stats(results: list[LibraryResult]) -> ReportStats:
    """
    Calcula estadísticas agregadas.
    Versión y vulnerabilidades se cuentan de forma independiente.
    """
    stats = ReportStats(total=len(results))
    for r in results:
        if r.version_status == "UPDATE": stats.outdated += 1
        elif r.version_status == "ERROR":  stats.errors += 1
        if r.vulnerabilities:
            stats.vulnerable += 1
    return stats


def generate_html_report(results: list[LibraryResult], output_path: str,
                         execution_time: float, logger: logging.Logger,
                         recent_vulns: list[RecentVuln] | None = None) -> str:
    """Genera el reporte HTML con Jinja2 y lo guarda en output_path."""
    stats = get_summary_stats(results)
    stats.recent_vulns = len(recent_vulns) if recent_vulns else 0
    now = datetime.now()

    # Vulnerables primero, luego actualizables, OK al final
    status_order = {"VULNERABLE": 0, "UPDATE": 1, "ERROR": 2, "OK": 3}
    sorted_results = sorted(results, key=lambda r: (status_order.get(r.status, 4), r.name.lower()))

    vulnerable_libs = [r for r in results if r.vulnerabilities]

    # Renderizar
    env = Environment(loader=FileSystemLoader(TEMPLATE_DIR), autoescape=True)
    template = env.get_template(TEMPLATE_FILE)

    html = template.render(
        report_date=now.strftime("%d/%m/%Y"),
        report_time=now.strftime("%H:%M:%S"),
        stats=stats,
        libraries=sorted_results,
        vulnerable_libs=vulnerable_libs,
        recent_vulns=recent_vulns or [],
        execution_time=f"{execution_time:.1f}",
    )

    # Guardar
    abs_path = os.path.abspath(output_path)
    try:
        with open(abs_path, "w", encoding="utf-8") as fh:
            fh.write(html)
    except (OSError, PermissionError):
        # Fallback si el archivo está bloqueado
        fallback = f"library_report_{now.strftime('%Y%m%d_%H%M%S')}.html"
        abs_path = os.path.abspath(fallback)
        with open(abs_path, "w", encoding="utf-8") as fh:
            fh.write(html)
        logger.warning("Archivo bloqueado, se guardó en: %s", abs_path)

    logger.info("Reporte HTML generado: %s", abs_path)
    return abs_path


# ============================================================================
# Main
# ============================================================================
def main() -> int:
    """Flujo principal: leer Excel → PyPI → OSV → HTML."""
    parser = argparse.ArgumentParser(
        description="Verificador de versiones y vulnerabilidades de librerías Python."
    )
    parser.add_argument("--input", "-i", required=True,
                        help="Ruta al archivo Excel con las librerías aprobadas.")
    parser.add_argument("--output", "-o", default="library_report.html",
                        help="Ruta de salida del reporte HTML (default: library_report.html).")
    args = parser.parse_args()

    logger = setup_logging()
    start_time = time.time()

    logger.info("=" * 60)
    logger.info("INICIO — Verificador de librerías Python")
    logger.info("=" * 60)

    try:
        # 1. Leer Excel
        libraries = read_approved_libraries(args.input, logger)

        # 2. Sesión HTTP
        session = create_http_session()

        # 3. Verificar versiones en PyPI
        logger.info("Verificando versiones en PyPI (%d librerías)...", len(libraries))
        results = check_all_versions(libraries, session, logger)

        # 4. Escanear vulnerabilidades de la versión aprobada en OSV.dev
        logger.info("Escaneando vulnerabilidades de versiones aprobadas en OSV.dev...")
        scan_vulnerabilities(libraries, results, session, logger)

        # 5. Buscar vulnerabilidades recientes (últimas 24h, cualquier versión)
        known_ids = set()
        for r in results:
            for v in r.vulnerabilities:
                known_ids.add(v.id)
        recent_vulns = scan_recent_vulnerabilities(libraries, known_ids, session, logger)

        # 6. Generar reporte
        elapsed = time.time() - start_time
        report_path = generate_html_report(results, args.output, elapsed, logger, recent_vulns)

        # Resumen en consola
        stats = get_summary_stats(results)
        stats.recent_vulns = len(recent_vulns)
        logger.info("-" * 60)
        logger.info("RESUMEN:")
        logger.info("  Total librerías:               %d", stats.total)
        logger.info("  Versión nueva disponible:      %d", stats.outdated)
        logger.info("  Con vulns (versión aprobada):   %d", stats.vulnerable)
        logger.info("  Vulns recientes (últimas 24h): %d", stats.recent_vulns)
        logger.info("  Errores de consulta:           %d", stats.errors)
        logger.info("-" * 60)
        logger.info("Reporte: %s", report_path)
        logger.info("Tiempo: %.1fs", elapsed)

        # Exit code 1 si hay vulnerabilidades (útil para CI/CD)
        return 1 if stats.vulnerable > 0 else 0

    except FileNotFoundError as exc:
        logger.error("Archivo no encontrado: %s", exc)
        return 2
    except ValueError as exc:
        logger.error("Error de validación: %s", exc)
        return 2
    except KeyboardInterrupt:
        logger.warning("Interrumpido por el usuario.")
        return 130
    except Exception as exc:
        logger.exception("Error inesperado: %s", exc)
        return 3


if __name__ == "__main__":
    # Asegurar UTF-8 en Windows
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8")
        except Exception:
            pass

    sys.exit(main())
